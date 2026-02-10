import openpyxl
from django.shortcuts import render
from django.db import transaction
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.response import Response
from .models import Client, PreSale, PreSaleItem
from apps.inventory.models import Product
from .serializers import ClientSerializer, PreSaleSerializer
from apps.accounting.models import Expense, AccountReceivable
from rest_framework import viewsets, permissions, status, serializers
from rest_framework.decorators import action
from openpyxl.styles import Font, Alignment, PatternFill

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import cm
from reportlab.lib.colors import HexColor
from reportlab.platypus import Table, TableStyle
from django.db.models import Sum, Count, F
from django.db.models.functions import TruncDate
from django.contrib.admin.models import LogEntry
from decimal import Decimal

class ClientViewSet(viewsets.ModelViewSet):
    serializer_class = ClientSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        user = self.request.user
        if user.role in ['GERENCIA', 'ALMACEN', 'DISTRIBUCION'] or user.is_staff:
            return Client.objects.all()
        return Client.objects.filter(seller=user)
    
    def perform_create(self, serializer):
        serializer.save(seller=self.request.user)
    
    @action(detail=True, methods=['post'])
    def capture_gps(self, request, pk=None):
        client = self.get_object()
        lat = request.data.get('latitude')
        lng = request.data.get('longitude')
        
        if not lat or not lng:
            return Response({"error": "Coordenadas requeridas"}, status=status.HTTP_400_BAD_REQUEST)
        
        client.latitude = lat
        client.longitude = lng
        client.gps_captured_by = request.user
        client.gps_captured_at = timezone.now()
        client.save()
        
        return Response({"message": "Ubicación GPS guardada", "latitude": lat, "longitude": lng})
    
    @action(detail=True, methods=['get'])
    def visit_history(self, request, pk=None):
        client = self.get_object()
        
        # Get all presales for this client
        presales = PreSale.objects.filter(client=client).select_related('seller', 'distributor').order_by('-created_at')
        
        history = []
        for ps in presales:
            visit_data = {
                "id": ps.id,
                "date": ps.created_at.strftime('%Y-%m-%d %H:%M'),
                "status": ps.status,
                "total_amount": float(ps.total_amount),
                "seller_name": ps.seller.get_full_name() if ps.seller else None,
                "distributor_name": ps.distributor.get_full_name() if ps.distributor else None,
                "sale_type": ps.sale_type
            }
            history.append(visit_data)
        
        return Response({
            "client": {
                "id": client.id,
                "business_name": client.business_name,
                "address": client.address,
                "latitude": str(client.latitude) if client.latitude else None,
                "longitude": str(client.longitude) if client.longitude else None
            },
            "visit_count": len(history),
            "visits": history
        })
    
    @action(detail=True, methods=['post'])
    def confirm_presale(self, request, pk=None):
        """Endpoint para que ALMACEN confirme que hay stock disponible"""
        if request.user.role not in ['ALMACEN', 'GERENCIA']:
            return Response(
                {"success": False, "message": "No tienes permisos para confirmar preventas"}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        presale = self.get_object()
        
        if presale.status != 'PENDIENTE':
            return Response(
                {"success": False, "message": "Solo se pueden confirmar preventas pendientes"}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        presale.status = 'CONFIRMADO'
        presale.save()
        
        return Response({
            "success": True, 
            "message": "Preventa confirmada correctamente",
            "data": PreSaleSerializer(presale).data
        })

class PreSaleViewSet(viewsets.ModelViewSet):
    serializer_class = PreSaleSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role in ['GERENCIA', 'ALMACEN', 'DISTRIBUCION'] or user.is_staff:
            queryset = PreSale.objects.select_related('client', 'seller', 'distributor').prefetch_related('items__product').all().order_by('-created_at')
        else:
            queryset = PreSale.objects.select_related('client', 'seller', 'distributor').prefetch_related('items__product').filter(seller=user).order_by('-created_at')
        
        status_param = self.request.query_params.get('status')
        if status_param:
            queryset = queryset.filter(status=status_param)
            
        return queryset

    def create(self, request, *args, **kwargs):
        """
        La lógica de descontar stock y crear items está en el 
        PreSaleSerializer.create() para mantener la atomicidad.
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        try:
            with transaction.atomic():
                presale = serializer.save(seller=self.request.user)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
        except serializers.ValidationError as e:
            return Response({"error": e.detail}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        
    # REPORTES ADICIONALES
    @action(detail=False, methods=['get'])
    def report_by_seller(self, request):
        start = request.query_params.get('start')
        end = request.query_params.get('end')
        
        data = PreSale.objects.filter(
            created_at__date__range=[start, end],
            status='CONFIRMADO'
        ).values('seller__username').annotate(
            total_sales=Sum('total_amount'),
            total_orders=Count('id')
        ).order_by('-total_sales')
        
        return Response(list(data))
    
    @action(detail=False, methods=['get'])
    def report_by_client(self, request):
        start = request.query_params.get('start')
        end = request.query_params.get('end')
        
        data = PreSale.objects.filter(
            created_at__date__range=[start, end],
            status='CONFIRMADO'
        ).values('client__business_name', 'client__id').annotate(
            total_purchases=Sum('total_amount'),
            total_orders=Count('id')
        ).order_by('-total_purchases')
        
        return Response(list(data))
    
    @action(detail=False, methods=['get'])
    def report_by_product(self, request):
        start = request.query_params.get('start')
        end = request.query_params.get('end')
        
        data = PreSaleItem.objects.filter(
            pre_sale__created_at__date__range=[start, end],
            pre_sale__status='CONFIRMADO'
        ).values('product__name', 'product__code').annotate(
            total_qty=Sum('quantity'),
            total_revenue=Sum(F('quantity') * F('price_at_sale'))
        ).order_by('-total_qty')
        
        return Response(list(data))

    # GENERAR EXCEL DE REPORTES DE VENTAS DE PRODUCTOS
    @action(detail=False, methods=['get'])
    def export_excel_report(self, request):
        start_date = request.query_params.get('start')
        end_date = request.query_params.get('end')

        report_data = PreSaleItem.objects.filter(
            pre_sale__created_at__date__range=[start_date, end_date],
            pre_sale__status='CONFIRMADO'
        ).values(
            'pre_sale__seller__username', 
            'product__name'
        ).annotate(
            total_qty=Sum('quantity'),
            total_revenue=Sum(F('quantity') * F('price_at_sale'))
        ).order_by('pre_sale__seller__username')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte por Vendedor"

        # Encabezados con estilo oscuro
        headers = ['VENDEDOR', 'PRODUCTO', 'CANTIDAD', 'TOTAL (Bs.)']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

        # Insertar los datos
        for item in report_data:
            ws.append([
                (item['pre_sale__seller__username'] or "S/V").upper(),
                item['product__name'].upper(),
                item['total_qty'],
                float(item['total_revenue'])
            ])
            ws.cell(row=ws.max_row, column=4).number_format = '#,##0.00 "Bs."'

        # Ajuste de anchos
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Ventas_{start_date}_al_{end_date}.xlsx"'
        wb.save(response)
        return response

    # GENERAR PDF DE NOTA DE ENTREGA STAR
    @action(detail=True, methods=['get'])
    def generate_pdf(self, request, pk=None):
        presale = self.get_object()
        items = presale.items.all()
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Nota_Entrega_{presale.id}.pdf"'

        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        red = HexColor('#A52A2A')

        # --- ENCABEZADO IZQUIERDO: Logo y contacto ---
        p.setFillColor(red)
        p.circle(3*cm, height - 3*cm, 1*cm, stroke=1, fill=0)
        p.setFont("Helvetica-Bold", 16)
        p.drawCentredString(3*cm, height - 3.2*cm, "STAR")
        
        p.setFillColor('black')
        p.setFont("Helvetica", 8)
        p.drawString(1.5*cm, height - 4.5*cm, "Av. Principal #123")
        p.drawString(1.5*cm, height - 5*cm, "Santa Cruz - Bolivia")
        p.drawString(1.5*cm, height - 5.5*cm, "Tel: 3-123456 / Cel: 70123456")

        # --- ENCABEZADO DERECHO: Título y folio ---
        p.setFillColor(red)
        p.setFont("Helvetica-Bold", 18)
        p.drawRightString(width - 2*cm, height - 2*cm, "NOTA DE ENTREGA")
        
        p.setFillColor('black')
        p.setFont("Helvetica-Bold", 10)
        p.drawRightString(width - 2*cm, height - 2.8*cm, f"N\u00ba {presale.id:06d}")
        
        # Fecha
        p.setFont("Helvetica", 8)
        fecha = presale.created_at
        p.drawRightString(width - 2*cm, height - 3.5*cm, f"DIA: {fecha.day:02d}  MES: {fecha.month:02d}  A\u00d1O: {fecha.year}")
        p.drawRightString(width - 2*cm, height - 4*cm, f"T/C: 6.96")

        # --- DATOS DEL CLIENTE ---
        y = height - 6*cm
        p.setFont("Helvetica-Bold", 9)
        p.drawString(2*cm, y, "Se\u00f1or(es):")
        p.setFont("Helvetica", 9)
        p.drawString(4.5*cm, y, presale.client.business_name.upper())
        p.line(4.5*cm, y - 0.1*cm, width - 2*cm, y - 0.1*cm)
        
        y -= 0.7*cm
        p.setFont("Helvetica-Bold", 9)
        p.drawString(2*cm, y, "Direcci\u00f3n:")
        p.setFont("Helvetica", 9)
        p.drawString(4.5*cm, y, presale.client.address or "")
        p.line(4.5*cm, y - 0.1*cm, width - 2*cm, y - 0.1*cm)
        
        y -= 0.7*cm
        p.setFont("Helvetica-Bold", 9)
        p.drawString(2*cm, y, "Tel:")
        p.setFont("Helvetica", 9)
        p.drawString(4.5*cm, y, presale.client.phone or "")
        p.line(4.5*cm, y - 0.1*cm, width - 2*cm, y - 0.1*cm)

        # --- TABLA DE PRODUCTOS ---
        y -= 1.5*cm
        table_data = [['ITEM', 'DETALLE', 'C/CAJA', 'C/UNIT', 'P/UNID', 'TOTAL']]
        
        # Convertir todo a USD
        for idx, item in enumerate(items, 1):
            precio_usd = float(item.price_at_sale) / 6.96
            subtotal_usd = float(item.subtotal) / 6.96
            table_data.append([
                str(idx),
                item.product.name.upper()[:35],
                '',
                str(item.quantity),
                f"${precio_usd:.2f}",
                f"${subtotal_usd:.2f}"
            ])
        
        # Rellenar filas vacías hasta 12
        while len(table_data) < 13:
            table_data.append(['', '', '', '', '', ''])
        
        table = Table(table_data, colWidths=[1.5*cm, 8*cm, 2*cm, 2*cm, 2.5*cm, 2.5*cm])
        table.setStyle(TableStyle([
            ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 9),
            ('FONT', (0, 1), (-1, -1), 'Helvetica', 8),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (5, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 0.5, 'black'),
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#f0f0f0')),
        ]))
        
        table.wrapOn(p, width, height)
        table.drawOn(p, 2*cm, y - 8*cm)

        # --- TOTAL ---
        y = y - 9*cm
        total_usd = float(presale.total_amount) / 6.96
        p.setFont("Helvetica-Bold", 11)
        p.drawString(width - 7*cm, y, "TOTAL $us:")
        p.drawRightString(width - 2*cm, y, f"${total_usd:.2f}")
        p.rect(width - 7.5*cm, y - 0.3*cm, 5*cm, 0.8*cm)

        # --- PIE DE PÁGINA ---
        y -= 2*cm
        p.setFont("Helvetica", 8)
        p.drawString(2*cm, y, "A cuenta: _______________________________________________")
        y -= 0.6*cm
        p.drawString(2*cm, y, "Son: _____________________________________________________")

        # --- FIRMAS ---
        y = 3*cm
        p.setFont("Helvetica-Bold", 9)
        p.drawString(3*cm, y, "RECIBI CONFORME")
        p.line(2.5*cm, y + 0.2*cm, 8*cm, y + 0.2*cm)
        
        p.drawString(width - 8*cm, y, "ENTREGUE CONFORME")
        p.line(width - 8.5*cm, y + 0.2*cm, width - 2*cm, y + 0.2*cm)
        
        distribuidor = presale.distributor.username.upper() if presale.distributor else "SIN ASIGNAR"
        p.setFont("Helvetica", 7)
        p.drawString(width - 8*cm, y - 0.5*cm, f"Distribuidor: {distribuidor}")

        p.showPage()
        p.save()
        return response
    
    @action(detail=False, methods=['get'])
    def dashboard_stats(self, request):
        # --- 1. CONTABILIDAD (Módulo 4) ---
        # Ventas Confirmadas (Ingresos Brutos)
        ingresos = PreSale.objects.filter(status='CONFIRMADO').aggregate(total=Sum('total_amount'))['total'] or 0
        # Egresos Totales (Nóminas + Pagos a Proveedores + Gastos varios)
        egresos = Expense.objects.aggregate(total=Sum('amount'))['total'] or 0
        # Deuda de Clientes (Cuentas por Cobrar reales)
        por_cobrar = AccountReceivable.objects.exclude(status='PAGADO').aggregate(total=Sum('remaining_balance'))['total'] or 0

        # --- 2. ALMACÉN (Módulo 2) ---
        # Usando stock_min de tu modelo Product
        productos_criticos = Product.objects.filter(
            stock__lte=F('stock_min'),
            is_active=True
        ).values('name', 'stock', 'stock_min').order_by('stock')[:5]

        # --- 3. SEGURIDAD (Módulo 6) ---
        # Trazabilidad real de Django
        acciones = LogEntry.objects.select_related('user').all().order_by('-action_time')[:5]
        logs_data = [{
            "time": log.action_time.strftime('%H:%M'),
            "user": log.user.username,
            "action": f"{log.get_action_flag_display()}: {log.object_repr}"
        } for log in acciones]

        # --- 4. VENTAS Y DISTRIBUCIÓN ---
        stats_entrega = PreSale.objects.values('status').annotate(count=Count('id'))
        top_productos = PreSaleItem.objects.values('product__name').annotate(
            total_qty=Sum('quantity')
        ).order_by('-total_qty')[:5]

        return Response({
            "ventas_totales": float(ingresos),
            "gastos_totales": float(egresos),
            "cuentas_por_cobrar": float(por_cobrar),
            "productos_criticos": list(productos_criticos),
            "recientes_logs": logs_data,
            "distribucion_estados": list(stats_entrega),
            "top_productos": list(top_productos)
        })
    
    # @action(detail=False, methods=['get'])
    # def dashboard_stats(self, request):
    #     # --- 1. Métricas de Ventas (Ya lo tenías) ---
    #     total_ventas = PreSale.objects.filter(status='CONFIRMADO').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    #     ventas_por_dia = PreSale.objects.filter(status='CONFIRMADO')\
    #         .annotate(dia=TruncDate('created_at'))\
    #         .values('dia')\
    #         .annotate(total=Sum('total_amount'))\
    #         .order_by('dia')

    #     # --- 2. Métricas de Distribución (Ya lo tenías) ---
    #     stats_entrega = PreSale.objects.values('status').annotate(count=Count('id'))
        
    #     # --- 3. Productos más vendidos (Ya lo tenías) ---
    #     top_productos = PreSaleItem.objects.values('product__name')\
    #         .annotate(total_qty=Sum('quantity'))\
    #         .order_by('-total_qty')[:5]

    #     # --- 4. MODIFICADO: Gastos y Utilidad (Módulo Contabilidad) ---
    #     # Sumamos todos los egresos registrados en el sistema
    #     total_gastos = Expense.objects.aggregate(Sum('amount'))['amount__sum'] or 0

    #     # --- 5. MODIFICADO: Stock Crítico (Módulo Almacén) ---
    #     # Buscamos productos donde la cantidad actual sea menor o igual al stock mínimo
    #     productos_criticos = Product.objects.filter(
    #         stock__lte=F('min_stock')
    #     ).values('name', 'stock', 'min_stock')[:5]

    #     # --- 6. MODIFICADO: Auditoría de Seguridad (Módulo Seguridad) ---
    #     # Aquí puedes usar LogEntry de Django o una lista simple de ejemplo si aún no tienes el modelo
    #     logs_recientes = [
    #         {"time": "Reciente", "user": "Sistema", "action": "Sincronización de inventario completada"},
    #         {"time": "Hoy", "user": request.user.username, "action": "Consulta de panel gerencial"}
    #     ]

    #     # --- RETORNO DE DATA ---
    #     return Response({
    #         "ventas_totales": total_ventas,
    #         "gastos_totales": total_gastos,         # NUEVO
    #         "productos_criticos": list(productos_criticos), # NUEVO
    #         "recientes_logs": logs_recientes,       # NUEVO
    #         "grafico_ventas": list(ventas_por_dia),
    #         "distribucion_estados": list(stats_entrega),
    #         "top_productos": list(top_productos)
    #     })