from rest_framework import viewsets, permissions, status
from rest_framework.response import Response
from .models import Product, Category, InventoryMovement
from .serializers import ProductSerializer, CategorySerializer, InventoryMovementSerializer
from rest_framework.decorators import action
from django.db.models import Sum, F
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill
import io

class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated]

class ProductViewSet(viewsets.ModelViewSet):
    serializer_class = ProductSerializer

    def get_queryset(self):
        if self.action == 'list':
            return Product.objects.select_related('category').filter(is_active=True)
        return Product.objects.select_related('category').all()

    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [permissions.IsAuthenticated()]
        
        if self.action in ['create', 'update', 'partial_update']:
            user = self.request.user
            if user.is_staff or (hasattr(user, 'role') and user.role in ['ALMACEN', 'gerente']):
                return [permissions.IsAuthenticated()]
            return [permissions.IsAdminUser()]
        return [permissions.IsAdminUser()]
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        products = Product.objects.select_related('category').filter(is_active=True)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        headers = ['CÓDIGO', 'NOMBRE', 'CATEGORÍA', 'UBICACIÓN', 'UNIDADES/CAJA', 'STOCK', 'STOCK MÍN', 'PRECIO COMPRA', 'P. HORIZONTAL', 'P. MAYORISTA', 'P. MODERNO']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        
        for p in products:
            ws.append([
                p.code,
                p.name,
                p.category.name if p.category else '',
                p.warehouse_location,
                p.units_per_box,
                p.stock,
                p.stock_min,
                float(p.purchase_price),
                float(p.price_horizontal),
                float(p.price_mayorista),
                float(p.price_moderno)
            ])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Productos.xlsx"'
        wb.save(response)
        return response
    
    @action(detail=False, methods=['post'])
    def import_excel(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No se proporcionó archivo'}, status=400)
        
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            created = 0
            updated = 0
            errors = []
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Si no hay código, saltar
                    continue
                
                try:
                    code, name, cat_name, location, units_box, stock, stock_min, p_compra, p_horiz, p_mayor, p_mod = row[:11]
                    
                    category, _ = Category.objects.get_or_create(name=cat_name)
                    
                    product, created_flag = Product.objects.update_or_create(
                        code=code,
                        defaults={
                            'name': name,
                            'category': category,
                            'warehouse_location': location or '',
                            'units_per_box': int(units_box or 1),
                            'stock': int(stock or 0),
                            'stock_min': int(stock_min or 5),
                            'purchase_price': float(p_compra or 0),
                            'price_horizontal': float(p_horiz or 0),
                            'price_mayorista': float(p_mayor or 0),
                            'price_moderno': float(p_mod or 0)
                        }
                    )
                    
                    if created_flag:
                        created += 1
                    else:
                        updated += 1
                        
                except Exception as e:
                    errors.append(f"Fila {row}: {str(e)}")
            
            return Response({
                'success': True,
                'created': created,
                'updated': updated,
                'errors': errors
            })
            
        except Exception as e:
            return Response({'error': str(e)}, status=400)
    
class InventoryMovementViewSet(viewsets.ModelViewSet):
    queryset = InventoryMovement.objects.select_related('product', 'user').all().order_by('-created_at')
    serializer_class = InventoryMovementSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def report(self, request):
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')

        # Filtrar solo egresos (ventas) que estén activos
        queryset = InventoryMovement.objects.filter(type='EGRESO')

        if start_date and end_date:
            # Filtramos por rango de fecha (formato YYYY-MM-DD)
            queryset = queryset.filter(created_at__date__range=[start_date, end_date])

        # Agrupamos por producto para obtener totales por ítem
        report_data = queryset.values(
            'product__id', 
            'product__name'
        ).annotate(
            total_qty=Sum('quantity'),
            # Calculamos el ingreso basado en el precio de venta (usando price_horizontal como referencia)
            total_revenue=Sum(F('quantity') * F('product__price_horizontal'))
        ).order_by('-total_qty')

        return Response(report_data)

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        product = serializer.validated_data['product']
        quantity = serializer.validated_data['quantity']
        mov_type = serializer.validated_data['type']

        # Lógica de actualización de Stock
        if mov_type == 'INGRESO':
            product.stock += quantity
        elif mov_type == 'EGRESO':
            if product.stock < quantity:
                return Response(
                    {"error": "Stock insuficiente para realizar el egreso"},
                    status=status.HTTP_400_BAD_REQUEST
                )
            product.stock -= quantity
        
        product.save()
        serializer.save(user=self.request.user)
        
        return Response(serializer.data, status=status.HTTP_201_CREATED)