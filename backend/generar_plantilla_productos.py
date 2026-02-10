import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Crear libro de Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Productos"

# Definir estilos
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Encabezados
headers = [
    'code',
    'name',
    'category',
    'description',
    'purchase_price',
    'price_horizontal',
    'stock',
    'stock_min',
    'units_per_box',
    'warehouse_location'
]

# Escribir encabezados
for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Ajustar anchos de columna
column_widths = {
    'A': 15,  # code
    'B': 30,  # name
    'C': 20,  # category
    'D': 40,  # description
    'E': 15,  # purchase_price
    'F': 15,  # price_horizontal
    'G': 10,  # stock
    'H': 10,  # stock_min
    'I': 15,  # units_per_box
    'J': 20   # warehouse_location
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Datos de ejemplo
ejemplos = [
    ['PROD001', 'Coca Cola 2L', 'Bebidas', 'Gaseosa Coca Cola 2 litros', 8.50, 12.00, 100, 20, 6, 'Estante A1'],
    ['PROD002', 'Aceite Fino 1L', 'Abarrotes', 'Aceite vegetal Fino 1 litro', 15.00, 20.00, 50, 10, 12, 'Estante B2'],
    ['PROD003', 'Arroz Grano de Oro 1kg', 'Abarrotes', 'Arroz blanco Grano de Oro 1kg', 6.00, 8.50, 200, 30, 20, 'Estante C1'],
    ['PROD004', 'Leche PIL 1L', 'Lácteos', 'Leche entera PIL 1 litro', 7.00, 10.00, 80, 15, 12, 'Refrigerador 1'],
    ['PROD005', 'Pan Blanco', 'Panadería', 'Pan blanco tradicional', 0.50, 1.00, 150, 50, 1, 'Vitrina A'],
]

# Escribir ejemplos
for row_idx, ejemplo in enumerate(ejemplos, start=2):
    for col_idx, valor in enumerate(ejemplo, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = valor
        cell.border = border
        cell.alignment = Alignment(horizontal='left', vertical='center')

# Agregar hoja de instrucciones
ws_instrucciones = wb.create_sheet("Instrucciones")
ws_instrucciones.column_dimensions['A'].width = 80

instrucciones = [
    "INSTRUCCIONES PARA IMPORTAR PRODUCTOS",
    "",
    "1. FORMATO DEL ARCHIVO:",
    "   - Debe ser un archivo Excel (.xlsx)",
    "   - La primera fila debe contener los encabezados exactos",
    "   - No cambies el nombre de las columnas",
    "",
    "2. COLUMNAS REQUERIDAS:",
    "   • code: Código único del producto (ej: PROD001)",
    "   • name: Nombre del producto",
    "   • category: Categoría del producto (se creará si no existe)",
    "   • description: Descripción del producto (opcional)",
    "   • purchase_price: Precio de compra en Bs",
    "   • price_horizontal: Precio de venta en Bs",
    "   • stock: Cantidad en inventario",
    "   • stock_min: Stock mínimo de alerta",
    "   • units_per_box: Unidades por caja (1 si se vende por unidad)",
    "   • warehouse_location: Ubicación en almacén (opcional)",
    "",
    "3. REGLAS IMPORTANTES:",
    "   - El código (code) debe ser único",
    "   - Los precios deben ser números positivos",
    "   - El stock debe ser un número entero",
    "   - Si el producto ya existe (mismo código), se actualizará",
    "",
    "4. CATEGORÍAS:",
    "   - Si la categoría no existe, se creará automáticamente",
    "   - Ejemplos: Bebidas, Abarrotes, Lácteos, Panadería, Limpieza",
    "",
    "5. VENTA POR CAJAS:",
    "   - Si units_per_box = 1: Se vende solo por unidad",
    "   - Si units_per_box > 1: Se puede vender por unidad o por caja",
    "   - Ejemplo: units_per_box = 6 significa 6 unidades por caja",
    "",
    "6. CÓMO IMPORTAR:",
    "   a) Completa la hoja 'Productos' con tus datos",
    "   b) Guarda el archivo",
    "   c) En el sistema, ve a Almacén → Inventario",
    "   d) Haz clic en 'Importar Excel'",
    "   e) Selecciona tu archivo",
    "   f) Revisa el resultado de la importación",
    "",
    "7. RESULTADO DE LA IMPORTACIÓN:",
    "   - Productos creados: Nuevos productos agregados",
    "   - Productos actualizados: Productos existentes modificados",
    "   - Errores: Productos con problemas (revisa el mensaje)",
    "",
    "8. CONSEJOS:",
    "   - Elimina las filas de ejemplo antes de importar",
    "   - Verifica que los precios sean correctos",
    "   - Usa códigos descriptivos (ej: BEB001, ABR001)",
    "   - Mantén una copia de respaldo de tu archivo",
]

for row_idx, texto in enumerate(instrucciones, start=1):
    cell = ws_instrucciones.cell(row=row_idx, column=1)
    cell.value = texto
    if "INSTRUCCIONES" in texto or texto.startswith(("1.", "2.", "3.", "4.", "5.", "6.", "7.", "8.")):
        cell.font = Font(bold=True, size=12)
    else:
        cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Guardar archivo
filename = 'plantilla_productos_importacion.xlsx'
wb.save(filename)
print(f"✅ Plantilla creada: {filename}")
print(f"📁 Ubicación: {filename}")
print("\nAbre el archivo, completa tus productos y súbelo al sistema.")
