"""
Script para generar plantilla Excel de ejemplo para importación de productos
Ejecutar: python generate_excel_template.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    
    # Headers
    headers = [
        'CÓDIGO', 'NOMBRE', 'CATEGORÍA', 'UBICACIÓN', 'UNIDADES/CAJA', 
        'STOCK', 'STOCK MÍN', 'PRECIO COMPRA', 'P. HORIZONTAL', 
        'P. MAYORISTA', 'P. MODERNO'
    ]
    
    ws.append(headers)
    
    # Estilo de headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Datos de ejemplo
    sample_data = [
        ['PROD001', 'Coca Cola 2L', 'Bebidas', 'A1-01', 12, 240, 50, 8.50, 12.00, 11.50, 11.00],
        ['PROD002', 'Arroz Grano de Oro 1kg', 'Abarrotes', 'B2-15', 20, 400, 100, 5.00, 7.50, 7.00, 6.80],
        ['PROD003', 'Aceite Fino 1L', 'Abarrotes', 'B2-20', 15, 180, 60, 12.00, 18.00, 17.00, 16.50],
        ['PROD004', 'Leche PIL 1L', 'Lácteos', 'C1-05', 24, 288, 72, 6.00, 9.00, 8.50, 8.20],
        ['PROD005', 'Pan Molde Ideal', 'Panadería', 'D3-10', 10, 150, 30, 4.50, 7.00, 6.50, 6.30],
    ]
    
    for row_data in sample_data:
        ws.append(row_data)
    
    # Ajustar ancho de columnas
    column_widths = [12, 25, 15, 12, 15, 10, 12, 15, 15, 15, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # Guardar archivo
    filename = 'Plantilla_Importacion_Productos.xlsx'
    wb.save(filename)
    print(f"✅ Plantilla creada exitosamente: {filename}")
    print(f"📊 Contiene {len(sample_data)} productos de ejemplo")
    print("\n📝 Instrucciones:")
    print("1. Edita el archivo Excel con tus productos")
    print("2. Mantén el formato de las columnas")
    print("3. CÓDIGO debe ser único para cada producto")
    print("4. UNIDADES/CAJA: cantidad de unidades que vienen en una caja (mínimo 1)")
    print("5. Importa el archivo desde el sistema")

if __name__ == "__main__":
    create_template()
